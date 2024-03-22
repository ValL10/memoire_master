from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages 
from django.contrib.auth.decorators import login_required
from .models import Profile
from django.http import Http404, HttpResponseForbidden, HttpResponse
import openpyxl


# Create your views here.
def index(request):
    return render(request, 'core/index.html')


def login_user(request):

    if request.method == 'POST':
        email = request.POST['useremail']
        password = request.POST['userpassword']

        user = authenticate(request, username=email, password=password)

        if user is not None:
            login(request, user)
            return redirect('scoreboard_index')
        
        else:
            return render(request, 'core/login.html', {
                'error' : "Invalid Credentials!"
            })
    else:
        return render(request, 'core/login.html')

def logout_user(request):
    logout(request)
    return redirect('index')


@login_required
def teacher_index(request):

    if request.user.profile.is_Admin is not True:
        return HttpResponseForbidden()
    
    enseignants = Profile.objects.filter(role_user=2).order_by('-created_at')
    return render(request, 'core/teachers/index.html', {'enseignants': enseignants, 'count': len(enseignants)})

@login_required
def search_teacher(request):
    if request.user.profile.is_Admin is not True:
        return HttpResponseForbidden()
    
    enseignants = Profile.objects.filter(role_user=2).order_by('-created_at')
    return render(request, 'core/teachers/index.html', {'enseignants': enseignants, 'count': len(enseignants)})

@login_required
def create_teacher(request):
    if request.user.profile.is_Admin is not True:
        return HttpResponseForbidden()        

    return render(request, 'core/teachers/create.html')

@login_required
def store_teacher(request):
    if request.user.profile.is_Admin is not True:
        return HttpResponseForbidden()
    
    required_teacher = ["first_name","last_name","last_name","email","password","matiere_enseigne"]

    if request.method == 'POST':
        file = request.FILES['teachersFile']

        datas_to_add = verifyFileUpload(file, required_teacher, 'create_teacher')
        
        result = Profile.StoreManyTeacher(datas_to_add)

        if result:
            message_ajout = 'Ajout reussie! Donnez aux personnes concernées leurs Nom d\'Utilisateur et leurs Mots de passes temporaires pour qu\'ils puissent les changer au plus vite'

            messages.success(request, message_ajout)
            return redirect('teachers_index')
         

    raise Http404("La page demandée n'a pas pu etre trouvéw")



@login_required
def students_index(request):

    if request.user.profile.is_Admin is not True:
        return HttpResponseForbidden()
    
    etudiants = Profile.objects.filter(role_user=1).order_by('-created_at')

    return render(request, 'core/students/index.html', {'etudiants': etudiants, 'count': len(etudiants)})

@login_required
def search_student(request):
    pass

@login_required
def create_student(request):
    if request.user.profile.is_Admin is not True:
        return HttpResponseForbidden()
    
    return render(request, 'core/students/create.html')

@login_required
def store_students(request):
    if request.user.profile.is_Admin is not True:
        return HttpResponseForbidden()
    
    required = ["first_name","last_name","ni_etudiant","username","email","password","niveau_etudiant","parcours_etudiant"]

    if request.method == 'POST':
        file = request.FILES['studentsFile']

        datas_to_add = verifyFileUpload(file, required, 'create_student')
        
        result = Profile.StoreManyStudents(datas_to_add)

        if result:
            message_ajout = 'Ajout reussie! Donnez aux personnes concernées leurs Nom d\'Utilisateur et leurs Mots de passes temporaires pour qu\'ils puissent les changer au plus vite'

            messages.success(request, message_ajout)
            return redirect('students_index')
         

    raise Http404("La page demandée n'a pas pu etre trouvéw")


# verification des fichiers uploader pour ajouter des utilissateurs

def verifyFileUpload(fichier, required, redirect_url):

    classeur = openpyxl.load_workbook(fichier)
    feuille = classeur.active

    donnees = []
            
    for ligne in feuille.iter_rows(values_only=True):
        donnee = []

        if len(ligne) != len(required):

            #print('Le fichier selectionné contient des erreurs', len(ligne), len(required))
            return redirect(redirect_url)
        
        for item in ligne:
            donnee.append(item)      
        
        donnees.append(donnee)

    return donnees


def verifyFileUpload(fichier, required, redirect_url):

    classeur = openpyxl.load_workbook(fichier)
    feuille = classeur.active

    donnees = []
            
    for ligne in feuille.iter_rows(values_only=True):
        donnee = []

        if len(ligne) != len(required):

            print('Le fichier selectionné contient des erreurs')
            return redirect(redirect_url)
        
        for item in ligne:
            if item is not None and item != '':
                donnee.append(item)      
        
        if(len(donnee) == len(required)):
            donnees.append(donnee)

    return donnees